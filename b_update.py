import io
import os
import string
from cmath import inf
import time
from datetime import date, timedelta
from math import ceil
from tkinter import Tk  # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename

import openpyxl as opyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd
import requests
from numpy import NaN
from openpyxl.utils.dataframe import dataframe_to_rows
from pandasql import sqldf


def interface():
    print("hahhaha you thought")

def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper().replace(' ','')) - ord('A')) + 1
    return num

def num2col(n):
    """Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA."""
    name = ''
    while n > 0:
        n, r = divmod (n - 1, 26)
        name = chr(r + ord('A')) + name
    return name

def covid_stats():
    #request csv's here
    states_url = "https://github.com/nytimes/covid-19-data/raw/master/rolling-averages/us-states.csv"
    states_download = requests.get(states_url).content

    counties_url = "https://raw.githubusercontent.com/nytimes/covid-19-data/master/rolling-averages/us-counties-2022.csv"
    counties_download = requests.get(counties_url).content

    #make two data frames here
    df_states = pd.read_csv(io.StringIO(states_download.decode('utf-8')))
    df_counties = pd.read_csv(io.StringIO(counties_download.decode('utf-8')))

    #after creating data frames fro each of the 
    #print(df_states.head)
    yday = date.today() - timedelta(days = 1)
    yday = str(yday)
    
    #queries and tables for yesterday
    q_states = f"SELECT date, state, cases_avg FROM df_states WHERE date = '{yday}' ORDER BY state"
    q_counties = f"SELECT date, county, state, cases_avg FROM df_counties WHERE date = '{yday}' ORDER BY state, county"
    yday_state_results = sqldf(q_states)
    yday_counties_results = sqldf(q_counties)

    #queries and tables for april 3
    q_bline_states = "SELECT date, state, cases_avg FROM df_states WHERE date = '2022-04-03' ORDER BY state"
    q_bline_counties = "SELECT date, county, state, cases_avg FROM df_counties WHERE date = '2022-04-03' ORDER BY state, county"
    bline_states = sqldf(q_bline_states)
    bline_counties = sqldf(q_bline_counties)

    #get percent diffrences for states
    state_percents = (yday_state_results['cases_avg'] / bline_states['cases_avg']) * 100
    yday_state_results['baseline 04/03/22'] = bline_states['cases_avg']
    yday_state_results['percent_change'] = state_percents


    #get percent diffrences for counties
    county_percents = (yday_counties_results['cases_avg'] / bline_counties['cases_avg']) * 100
    yday_counties_results['baseline 04/03/22'] = bline_counties['cases_avg']
    yday_counties_results['percent_change'] = county_percents

    return yday_state_results, yday_counties_results

#modify to grow horizontally
def update_inter(fname, yday_state_results, yday_counties_results):
    #premade file vs file creation 
    # inter = opyxl.load_workbook(filename = fname)
    # states_data = inter["states"]
    # counties_data = inter["counties"]
    inter = Workbook()
    states_data = inter.create_sheet("states")
    counties_data = inter.create_sheet("counties")
    

    for r in dataframe_to_rows(yday_state_results, index = False, header = True):
        states_data.append(r)
    
    for r in dataframe_to_rows(yday_counties_results, index = False, header = True):
        counties_data.append(r)

    #ciel all of the values in column F + 3259 for counties and e + 57 for states
    # state_cell = states_data["E" + str(row)].value
    # county_cell = counties_data["F" + str(row)].value
    for row in range(2, 3259):
        if counties_data["F" + str(row)].value > 0 and counties_data["F" + str(row)].value < inf:
            counties_data["F" + str(row)].value = ceil(counties_data["F" + str(row)].value)
    
        else:
            counties_data["F" + str(row)].value = ceil(counties_data["D" + str(row)].value) * 100
            
    for row in range(2, 58):
       states_data["E" + str(row)].value = ceil(states_data["E" + str(row)].value)


    des = fname + "TO BE DELETED COVID_STATS.xlsx"
    inter.save(des)

#col E is county percent
#col H is state percent
#county is C
#statename is H
#C6E0B4 col e
#FFE699 col H

def apply_to_covid(inter, covid_letter):
    src = opyxl.load_workbook(filename=inter)
    des = opyxl.load_workbook(filename=covid_letter)
    db = des.active
    county_data = src['counties']
    states_data = src['states']
    
    #fill state and county at the same time
    des_row = 2
    county = db["D" + str(des_row)].value
    state = db["I" + str(des_row)].value
    state = state.upper().replace(' ','')
    county = county.upper().replace(' ','')   

    #one massive while that will fill the table
    #arrbitraryly choosing county as index
    
    not_found = []

    while county != None:
        
        #state operation 
        state_row = 2

        sindex = states_data["B" + str(state_row)].value
        sindex = sindex.upper().replace(' ','')

        #linear search of src data 
        while sindex != None and sindex != state:
            state_row += 1
            sindex = states_data["B" + str(state_row)].value
            if sindex != None:
                sindex = sindex.upper().replace(' ','')

        if sindex == None:
            print(f"not found: {state}")
            not_found.append(state, "<-state")
        else:
            db["H" + str(des_row)].value = (states_data["E" + str(state_row)].value)
            


        #county operation
        county_row = 2
        cindex = county_data["B" + str(county_row)].value
        sindex = county_data["C" + str(county_row)].value
        cindex = cindex.upper().replace(' ','')
        sindex = sindex.upper().replace(' ','')

        exist_index = county + state
        com_index = cindex + sindex

        #linear search of src data for county
        while cindex != None and com_index != exist_index:
            county_row += 1
            cindex = county_data["B" + str(county_row)].value
            sindex = county_data["C" + str(county_row)].value
            if cindex != None:
                cindex = cindex.upper().replace(' ','')
                sindex = sindex.upper().replace(' ','')
                com_index = cindex + sindex

        if cindex == None:
            print(f"county not found: {county}, {state}")
            not_found.append(county, "<-county", state)
        else:
            db["E" + str(des_row)].value = (county_data["F" + str(county_row)].value)
            

        des_row += 1
        county = db["D" + str(des_row)].value
        state = db["I" + str(des_row)].value
        if county != None: 
            state = state.upper().replace(' ','')
            county = county.upper().replace(' ','')


    color(db)
    
    des.save(covid_letter)

#coloring
def color(db):

    r = 2
    currentCell = db["E" + str(r)].value

    while currentCell != None:
        db.cell(column = col2num("E"), row = r).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        db.cell(column = col2num("H"), row = r).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        r += 1
        currentCell = db["E" + str(r)].value

#driver down here
# print("getting covid_stats")
# x,y = covid_stats()

# # print("now updating inter.xlsx")
# update_inter('inter.xlsx',x ,y)
# print("done updating")
# print("applying changes to covid spreadsheet")
# apply_to_covid('inter.xlsx','COVID COUNTY v1.xlsx')

# fdir = input("what is file ddfoasjdlfkjaslkdfj")

if __name__ == "__main__":
    print("please select the covid county file ")
    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    county_file = askopenfilename() # show an "Open" dialog box and return the path to the selected file
    
    ndex = 0
    for pos, i in enumerate(county_file):
        if i == "/":
            index = pos
    fname = county_file[:index + 1]
    des = fname + "TO BE DELETED COVID_STATS.xlsx"

    print("getting covid stats .....")
    x, y = covid_stats()
    print("creating intermediate file")
    update_inter(fname , x, y)
    print("applying updates ... please wait")
    apply_to_covid(des, county_file)
    print("file updated")
    print("Please REMOVE 'TO BE DELETED COVID STATS.xlsx'")
    time.sleep(5)
